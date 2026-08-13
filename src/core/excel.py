"""Excel 산출물.

시트 하나에 표가 하나일 수도, 여럿일 수도 있다(처리요약처럼).
표를 만드는 쪽은 여기에 컬럼과 행만 넘기고, 서식은 전부 이 모듈이 책임진다.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Final, Iterable, Sequence

__all__ = ["Block", "ExcelError", "Sheet", "write_workbook"]


class ExcelError(Exception):
    """Excel 을 쓰지 못했을 때."""


#: 시트 이름에 쓸 수 없는 문자 (Excel 규칙).
_FORBIDDEN: Final[str] = r"[]:*?/\\"
_MAX_SHEET_NAME: Final[int] = 31

#: 컬럼 폭 상한. 원본 레코드 JSON 같은 긴 값이 화면을 다 먹지 않게 한다.
_MAX_WIDTH: Final[int] = 60
_MIN_WIDTH: Final[int] = 8

#: 심각도별 배경색. 눈으로 먼저 걸러야 하는 표라서 색을 넣는다.
_SEVERITY_FILL: Final[dict[str, str]] = {
    "높음": "FFC7CE",
    "보통": "FFEB9C",
    "낮음": "E2EFDA",
}


@dataclass(slots=True)
class Block:
    """시트 안의 표 하나."""

    columns: Sequence[str]
    rows: Sequence[dict[str, Any]]
    title: str | None = None


@dataclass(slots=True)
class Sheet:
    name: str
    blocks: list[Block] = field(default_factory=list)


def _safe_name(name: str, used: set[str]) -> str:
    cleaned = "".join(" " if ch in _FORBIDDEN else ch for ch in name).strip()
    cleaned = (cleaned or "시트")[:_MAX_SHEET_NAME]
    candidate, suffix = cleaned, 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = cleaned[: _MAX_SHEET_NAME - len(tail)] + tail
        suffix += 1
    used.add(candidate)
    return candidate


def _cell_value(value: Any) -> Any:
    """Excel 에 넣을 값으로 다듬는다.

    '=' 로 시작하는 문자열은 Excel 이 수식으로 해석한다.
    감사추적 값이 수식이 되어 다른 값으로 표시되면 안 되므로 그대로 보이게 막는다.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float, datetime)):
        return value
    text = str(value)
    if text[:1] in ("=", "+", "@"):
        return f"'{text}"
    return text


def _width(column: str, values: Iterable[Any]) -> int:
    longest = max((len(str(v)) for v in values), default=0)
    return max(_MIN_WIDTH, min(_MAX_WIDTH, max(longest, len(column)) + 2))


def write_workbook(sheets: Sequence[Sheet], path: Path | str) -> Path:
    """시트들을 .xlsx 로 쓴다. 쓴 경로를 반환."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - 설치 안내용
        raise ExcelError(
            "Excel 로 쓰려면 openpyxl 이 필요합니다: pip install openpyxl"
        ) from exc

    if not sheets:
        raise ExcelError("쓸 시트가 없습니다.")

    target = Path(path)
    if target.parent and not target.parent.exists():
        target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="44546A")
    title_font = Font(bold=True, size=12)
    top = Alignment(vertical="top", wrap_text=False)

    used: set[str] = set()
    for sheet in sheets:
        worksheet = workbook.create_sheet(_safe_name(sheet.name, used))
        row_index = 1
        first_header_row: int | None = None
        widths: dict[int, int] = {}

        for block in sheet.blocks:
            if block.title:
                cell = worksheet.cell(row=row_index, column=1, value=block.title)
                cell.font = title_font
                row_index += 2

            header_row = row_index
            if first_header_row is None:
                first_header_row = header_row
            for column_index, name in enumerate(block.columns, start=1):
                cell = worksheet.cell(row=header_row, column=column_index, value=name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = top
            row_index += 1

            severity_index = (
                list(block.columns).index("심각도") + 1
                if "심각도" in block.columns
                else None
            )

            for record in block.rows:
                for column_index, name in enumerate(block.columns, start=1):
                    cell = worksheet.cell(
                        row=row_index,
                        column=column_index,
                        value=_cell_value(record.get(name)),
                    )
                    cell.alignment = top
                if severity_index is not None:
                    color = _SEVERITY_FILL.get(str(record.get("심각도", "")))
                    if color:
                        worksheet.cell(row=row_index, column=severity_index).fill = (
                            PatternFill("solid", fgColor=color)
                        )
                row_index += 1

            for column_index, name in enumerate(block.columns, start=1):
                candidate = _width(name, (r.get(name) for r in block.rows))
                widths[column_index] = max(widths.get(column_index, 0), candidate)
            row_index += 1        # 표 사이 한 줄 띄우기

        for column_index, width in widths.items():
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        # 표가 하나뿐인 시트만 고정·필터를 건다. 여러 표가 섞이면 오히려 방해된다.
        if len(sheet.blocks) == 1 and first_header_row is not None:
            block = sheet.blocks[0]
            worksheet.freeze_panes = worksheet.cell(row=first_header_row + 1, column=1)
            if block.rows:
                last = get_column_letter(len(block.columns))
                worksheet.auto_filter.ref = (
                    f"A{first_header_row}:{last}{first_header_row + len(block.rows)}"
                )

    workbook.save(target)
    return target
